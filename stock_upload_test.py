
# -*- coding: utf-8 -*-
# encoding = utf8 

import sys
import openpyxl
import datetime
import traceback
import string

from datetime 						import date
from openpyxl						import Workbook
from openpyxl						import load_workbook
from openpyxl.styles				import Color, PatternFill, Font, Border, Side, Alignment, Protection, colors
from openpyxl.comments  			import Comment
from openpyxl.worksheet				import Worksheet
from openpyxl 						import *
from openpyxl.utils 				import get_column_letter
from openpyxl.worksheet.table 		import Table, TableStyleInfo

#- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -

today 				= '20180101'
last_export_day 	= '20171231'

reserv_list 			= []
stock_list				= []
sub_onec_list			= []
onec_list				= []
last_itd_stock_list		= []
last_itd_cross_list		= []
new_itd_stock_line		= []
fsasoc_stock_list		= []

path 			= '-'
reserv_file 	= '{path}\\sap_exports\\reservations_{today}.xlsx'.format(path=path, today=today)
stock_file 		= '{path}\\sap_exports\\sq00_{today}.xlsx'.format(path=path, today=today)
last_itd_file	= '{path}\\last_files\\Stock_Update_{day}.xlsx'.format(path= path, day=last_export_day)
onec_file		= '{path}\\1c_exports\\1C_{day}.xlsx'.format(path= path, day=today)

#- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
#- - - - - H E L I O S   L I S T - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
#- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -

helios_list = [
{'helios_code':'I','business_line':'IND',},
]

#- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
#- - - - - E L O G - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
#- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -

def elog( name, msg ): 
	
		''' creates log file with Args:
		name 	- string - name of log file
		msg 	- string - text of error message
		mem_loc - string - location in memory for saving log_file
		'''
		with open( '{mem_loc}\\elog_{name}.txt'.format(name=name,mem_loc='{}\\logs\\'.format(path)), "a") as log_file:
			log_file.write( '\n{sep}'.format( sep = '------------------------------------------------------' ) )
			log_file.write( '\n{msg}'.format( msg = msg ) )
			log_file.write( '\n{sep}'.format( sep = '------------------------------------------------------' ) )

#- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
#- - - - - G E T   S T A T U S  R U  N A M E - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
#- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
			
def get_ru_status_name( status_id ):
	
	try:
		if status_id == '30' 	: status_ru_nme = u'Доступен к закупке и продаже'
		if status_id == '31' 	: status_ru_nme = u'Продажа из наличия, заблокирован к закупке'
		if status_id == '99' 	: status_ru_nme = u'Заблокирован для закупки и продаж'
		if status_id == '0' 	: status_ru_nme = u'Не установлен'
	
		return status_ru_nme
	
	except:
		return ''	

#- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
#- - - - - G E T   S T A T U S  E N  N A M E - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
#- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -

def get_en_status_name( status_id ):

	try:
		if status_id == '30' 	: status_en_nme = 'Open for sales and purchasing'
		if status_id == '31' 	: status_en_nme = 'Open for sales only'
		if status_id == '99' 	: status_en_nme = 'Blocked for sales and purchasing'
		if status_id == '0' 	: status_en_nme = 'Initialisation'

		return status_en_nme

	except:
		return ''

#- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
#- - - - - G E T   B U S I N E S S   L I N E - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
#- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -

def get_bu_name( helios_code ):

	try:
		bu_name = '-'
		for line in helios_list:
			if helios_code.lower() == line['helios_code'].lower():
				bu_name = line['business_line']

		return bu_name

	except Exception as e:
		elog('get_bu_name', '{d}_{type}\n{trb1}\n{trb2} '.format(d=datetime.datetime.now(),
																	type=str(type(e)),
																	trb1=traceback.format_tb(sys.exc_info()[2])[0],
																	trb2=str(sys.exc_info()[1])))

		return '-'

#- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
#- - - - - G E T   S T A T U S   I D - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
#- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
			
def get_status_id( status_nme ):
	
	try:
		if status_nme.lower() == 'Доступен к закупке и продаже'.lower() 				: status_id = '30'
		if status_nme.lower() == 'Продажа из наличия, заблокирован к закупке'.lower() 	: status_id = '31'
		if status_nme.lower() == 'Заблокирован для закупки и продаж'.lower() 			: status_id = '99'
		if status_nme.lower() == 'Не установлен'.lower() 								: status_id = '0'
		
		return status_id
		
	except:
		return ''

#- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
#- - - - - R E S E R V A T I O N   L O A D S - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
#- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -

def get_by_type(type_nme):

	try:
		if type_nme.lower() == 'ЗАБАЛАНС'.lower() 		: status = 'out of stock'
		if type_nme.lower() == 'REF'.lower() 			: status = 'refurbished'
		if type_nme.lower() == 'NEW'.lower() 			: status = 'out of stock'
		if type_nme.lower() == ''.lower() 				: status = 'out of stock'
		if type_nme.lower() == 'Return'.lower() 		: status = 'out of stock'
		if type_nme.lower() == 'Scrap'.lower() 			: status = 'delete'

		return status

	except:
		return ''

#- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
#- - - - - R E S E R V A T I O N   L O A D S - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
#- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -

def reserv_loading():
	
	try:
		#print ('reserv_loading start time :', datetime.datetime.now())
		reserv_sub_list 		= []
		reserv_materials_list 	= []
		
		material_str			= 'Material'	.lower()
		qty_str					= 'Diff. qty'	.lower()
		reserv_nme_str 			= 'SLoc'		.lower()
		
		wb_reserv 				= load_workbook(reserv_file)
		ws 						= wb_reserv.active		#initiating source sheet
		
		title_row = 2

		# looking for start row number
		for row_index in range(1,20):
			for col_index in range(1,50):
				if str(ws.cell(row=row_index,column=col_index).value).lower() == 'Material'.lower() :
					title_row=row_index	
		
		for col_index in range(1,20):
			cur_cll = str(ws.cell( row=title_row,column=col_index).value).lower()
			if material_str 		in cur_cll	:	material_col 	= col_index
			if qty_str 				in cur_cll	:	qty_col 		= col_index
			if reserv_nme_str 		in cur_cll	:	reserv_nme_col 	= col_index

		#looking for max row number
		for row_index in range(1,ws.max_row + 1):
			if str( ws.cell( row=row_index,column=2).value).lower() == '*':
				sht_max_row=row_index - 2	
			
		for row_index in range(title_row + 2, sht_max_row + 1 ):
			#creating list
			if ( ws.cell( row=row_index,column=material_col).value is not None ):
			
				reserv_sub_line = {}

				reserv_sub_line['material'		] 	= str(ws.cell(row=row_index,column=material_col	).value).strip()
				reserv_sub_line['qty'			] 	= str(ws.cell(row=row_index,column=qty_col		).value)
				reserv_sub_line['reserv_nme_str'] 	= str(ws.cell(row=row_index,column=reserv_nme_col).value).strip()
				if (reserv_sub_line['reserv_nme_str'] == 'None') or (reserv_sub_line['reserv_nme_str'] is None ):
					reserv_sub_line['reserv_nme_str'] = '4000'
				
				reserv_sub_list.append(reserv_sub_line)
		
		for line in reserv_sub_list :
			reserv_materials_list.append(line['material'])
		reserv_materials_list = set(reserv_materials_list)
		
		for material in reserv_materials_list:
			
			reserv_line					= {}
			reserv_line['material']		= material
			reserv_line['res_4000_qty']	= 0
			reserv_line['res_3400_qty']	= 0
			
			for line in reserv_sub_list :
				if (line['material'] == material ) and (line['reserv_nme_str'] == '4000') :
					reserv_line['res_4000_qty']	+= float(line['qty'])
				if (line['material'] == material ) and (line['reserv_nme_str'] == '3400') :
					reserv_line['res_3400_qty']	+= float(line['qty'])
			
			reserv_list.append(reserv_line)
		
		#for line in reserv_list:
			#print (line)
		#print ('reserv_loading end time :', datetime.datetime.now())
		
		return reserv_list

	except Exception as e:
		elog('reserv_loading','{d}_{type}\n{trb1}\n{trb2} '.format( d = datetime.datetime.now(),
																  type = str(type(e)),
																  trb1 = traceback.format_tb( sys.exc_info()[2] )[0],
																  trb2 = str( sys.exc_info()[1] ) ))
		return []
		
#- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - 
#- - - - - S T O C K   L O A D S - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - 
#- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - 

def stock_loading():
	
	try:
		#print ('stock_loading start time :', datetime.datetime.now() )
		
		material_str		= 'Material'	.lower()
		status_str			= 'MS'			.lower()
		qty_4000_str 		= 'SoH 4000'	.lower()
		qty_3400_str 		= 'SoH 3400'	.lower()
		description_str 	= 'Description'	.lower()
		helios_str 			= 'Helios'		.lower()
		pdt_str 			= 'PDT'			.lower()

		wb_stock	= load_workbook(stock_file)
		ws 			= wb_stock.active #initiating source sheet
		
		#looking for max row number
		for row_index in range(1,ws.max_row + 1):
			if str( ws.cell( row=row_index,column=2).value).lower() == '*':
				sht_max_row=row_index - 2
		
		#print ('max row stock file :', sht_max_row	)
		
		title_row=5
		
		for row_index in range(1,20):
			for col_index in range(1,50):
				if str(ws.cell(row=row_index,column=col_index).value).lower() == 'Material'.lower() :
					title_row=row_index

		for col_index in range(1,50):
			cur_cll = str(ws.cell( row=title_row,column=col_index).value).lower()
			if material_str 		== cur_cll	:	material_col 	= col_index
			if status_str 			in cur_cll	:	status_col 		= col_index
			if qty_4000_str 		in cur_cll	:	qty_4000_col 	= col_index
			if qty_3400_str 		in cur_cll	:	qty_3400_col 	= col_index
			if description_str 		in cur_cll	:	description_col = col_index
			if helios_str 			in cur_cll	:	helios_col 		= col_index
			if pdt_str 				in cur_cll	:	pdt_col 		= col_index

		#print (title_row)
		#print (sht_max_row)

		#for row_index in range(title_row + 2, 2000 ):
		for row_index in range(title_row + 2, sht_max_row + 1 ):
			#creating list
			if ( ws.cell( row=row_index,column=material_col).value is not None ):
			
				stock_line = {}
				try:
					stock_line['material'] = str(ws.cell(row=row_index,column=material_col).value).strip()
				except:
					stock_line['material'] = ''
					print ('Stock ERROR - material',row_index)

				try:
					stock_line['status'	] = str(ws.cell(row=row_index,column=status_col).value)
				except:
					stock_line['status'] = ''
					print ('Stock ERROR - status', row_index)

				try:
					stock_line['qty_4000'] = float(ws.cell(row=row_index,column=qty_4000_col).value)
				except:
					try:
						str_4000 = (str(ws.cell(row=row_index, column=qty_4000_col).value)).replace('.', '')
						str_4000 = str_4000.replace(',', '.')
						stock_line['qty_4000'] = float(str_4000)
					except:
						stock_line['qty_4000'] = 0.0
						print ('Stock ERROR - qty_4000', row_index)

				try:
					stock_line['qty_3400'] = float(ws.cell(row=row_index,column=qty_3400_col).value)
				except:
					stock_line['qty_3400'] = 0.0
					print ('Stock ERROR - qty_3400', row_index)

				try:
					stock_line['description'] = str(ws.cell(row=row_index,column=description_col).value)
				except:
					stock_line['description'] = ''
					print ('Stock ERROR - description', row_index)

				try:
					stock_line['helios_code'] = str(ws.cell(row=row_index,column=helios_col).value)
				except:
					stock_line['helios_code'] = ''
					print ('Stock ERROR - helios_code', row_index)

				try:
					stock_line['pdt'] = int(ws.cell(row=row_index,column=pdt_col).value)
				except:
					stock_line['pdt'] = '-'
					print ('Stock ERROR - pdt', row_index)

				stock_list.append(stock_line)

		return stock_list
	
	except Exception as e:
		elog('stock_loading','{d}_{type}\n{trb1}\n{trb2} '.format( d = datetime.datetime.now(),
																 type = str(type(e)),
																 trb1 = traceback.format_tb( sys.exc_info()[2] )[0],
																 trb2 = str( sys.exc_info()[1] ) ))
		return []

#- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
#- - - - - 1C  F I L E   L O A D S - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
#- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -

def onec_loading():

	try:

		material_str	= 'Номенклатура.Артикул'.lower()
		unit_str		= 'Номенклатура.Единица хранения'.lower()
		type_str 		= 'Ячейка.Качество'.lower()
		qty_str 		= 'Количество'.lower()


		wb_stock	= load_workbook(onec_file)
		ws 			= wb_stock.active #initiating source sheet

		#looking for max row number
		for row_index in range(1,ws.max_row + 1):
			if str( ws.cell(row=row_index,column=1).value).lower() == 'Итого'.lower():
				sht_max_row=row_index - 1

		#print ('max row stock file :', sht_max_row)

		title_row=5

		for row_index in range(1,10):
			for col_index in range(1,10):
				if material_str in str(ws.cell(row=row_index,column=col_index).value).lower():
					title_row=row_index

		for col_index in range(1,50):
			cur_cll = str(ws.cell( row=title_row,column=col_index).value).lower()
			if material_str 	in cur_cll	:	material_col 	= col_index
			#if unit_str 		in cur_cll	:	unit_col 		= col_index
			if type_str 		in cur_cll	:	type_col		= col_index
			if qty_str 			in cur_cll	:	qty_col 		= col_index

		#for row_index in range(title_row + 2, 2000 ):
		for row_index in range(title_row + 2, sht_max_row + 1 ):
			#creating list
			if ws.cell( row=row_index,column=material_col).value is not None:

				sub_onec_line = {}
				try:
					sub_onec_line['material'] = str(ws.cell(row=row_index,column=material_col).value).strip()
				except:
					sub_onec_line['material'] = ''
					print ('1C ERROR',row_index)

				'''
				try:
					sub_onec_line['unit'] = ws.cell(row=row_index,column=unit_col).value
				except:
					sub_onec_line['unit'] = ''
					print ('1C ERROR', row_index)
				'''

				try:
					tpe = ws.cell(row=row_index,column=type_col).value
				except:
					tpe = ''
					print ('1C ERROR', row_index)
				if tpe is None: tpe = ''
				sub_onec_line['status'] = get_by_type(tpe)

				try:
					sub_onec_line['qty'] = float(ws.cell(row=row_index,column=qty_col).value)
				except:
					sub_onec_line['qty'] = 0.0
					print ('1C ERROR', row_index)

				sub_onec_list.append(sub_onec_line)

		for material in sub_onec_list:
			onec_line = {}
			indicator = 0
			for line in onec_list:
				if (material['material'] == line['material']) and (material['status']==line['status']):
					line['qty'] += material['qty']
					indicator = 1
			if indicator == 0:
				onec_line['material'] = material['material']
				onec_line['status'] = material['status']
				onec_line['qty'] = material['qty']
				onec_list.append(onec_line)

		#for line in onec_list:
			#print (line)

		return onec_list

	except Exception as e:
		elog('onec_loading','{d}_{type}\n{trb1}\n{trb2} '.format( d = datetime.datetime.now(),
																 type = str(type(e)),
																 trb1 = traceback.format_tb( sys.exc_info()[2] )[0],
																 trb2 = str( sys.exc_info()[1] ) ))
		return []

#- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
#- - - - - L A S T   I T D   L O A D S - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
#- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -

def last_itd_loading():
	
	try:
		material_str		= 'Material'
		description_str		= 'Description'
		status_str			= 'Material Status'
		stock_4000_str		= 'Free Stock'
		stock_3400_str		= 'Contract Stock'
		bu_str				= 'Business Unit'
		pdt_str				= 'Delivery time'

		material_1_str		= 'Material'
		material_2_str		= 'Material2'
		material_3_str		= 'Material3'
		material_4_str		= 'Material4'

		wb_last_itd			= load_workbook(last_itd_file)
		ws_1				= wb_last_itd['Stock']#initiating source sheet
		ws_2				= wb_last_itd['Cross reference ITD']#initiating source sheet

		# - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
		#WS 2
		# - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
		#looking for max row number
		for row_index in range(1,ws_2.max_row + 1):
			if str( ws_2.cell( row=row_index,column=1).value) != 'None':
				sht_max_row = row_index

		#print ('max row last itd file :', sht_max_row)

		for col_index in range(1,30):
			cur_cll = str(ws_2.cell( row=1,column=col_index).value)
			if material_1_str 		== cur_cll		:	material_1_col 		= col_index
			if material_2_str 		== cur_cll		:	material_2_col 		= col_index
			if material_3_str 		== cur_cll		:	material_3_col 		= col_index
			if material_4_str 		== cur_cll		:	material_4_col 		= col_index

		for row_index in range(2, sht_max_row + 1):
			#creating list

			last_itd_cross_line = {}

			last_itd_cross_line['material1'	] = str(ws_2.cell(row=row_index ,column=material_1_col).value).strip()
			last_itd_cross_line['lenth'] = 1
			#Material 2
			if (str(ws_2.cell(row=row_index, column=material_2_col).value) != 'None') :
				last_itd_cross_line['material2'	] = str(ws_2.cell(row=row_index ,column=material_2_col).value).strip()
				last_itd_cross_line['lenth'] += 1
			else:
				last_itd_cross_line['material2'] = ''

			#Material 3
			if (str(ws_2.cell(row=row_index, column=material_3_col).value) != 'None'):
				last_itd_cross_line['material3'	] = str(ws_2.cell(row=row_index ,column=material_3_col).value).strip()
				last_itd_cross_line['lenth'] += 1
			else:
				last_itd_cross_line['material3'] = ''

			#Material 4
			if (str(ws_2.cell(row=row_index, column=material_4_col).value) != 'None'):
				last_itd_cross_line['material4'	] = str(ws_2.cell(row=row_index ,column=material_4_col).value).strip()
				last_itd_cross_line['lenth'] += 1
			else:
				last_itd_cross_line['material4'] = ''

			last_itd_cross_list.append(last_itd_cross_line)

		#for line in last_itd_list:
			#print (line)
		#print ('last_itd_loading end time :', datetime.datetime.now() )
		
		return last_itd_stock_list, last_itd_cross_list
	
	except Exception as e:
		elog('last_itd_loading','{d}_{type}\n{trb1}\n{trb2} '.format( d=datetime.datetime.now(),
																	type=str(type(e)),
																	trb1=traceback.format_tb( sys.exc_info()[2] )[0],
																	trb2=str( sys.exc_info()[1] ) ))
		return [], []
		
#- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
#- - - - - N E W   S T O C K    L I S T- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
#- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -

def gen_new_stock_list():
	
	try:
		new_itd_stock_list = []
		for stock_line in stock_list:
			new_itd_stock_line = {}
			new_itd_stock_line['material'		] 	= stock_line['material']
			new_itd_stock_line['status_id'		] 	= stock_line['status']
			new_itd_stock_line['description'	] 	= stock_line['description']
			new_itd_stock_line['pdt'			] 	= stock_line['pdt']
			new_itd_stock_line['status_name'	] 	= get_ru_status_name(stock_line['status'])
			new_itd_stock_line['stock_4000'		] 	= stock_line['qty_4000']
			new_itd_stock_line['stock_3400'		] 	= stock_line['qty_3400']

			bu_name = get_bu_name(stock_line['helios_code'])
			#print (bu_name)
			if bu_name != '-':
				new_itd_stock_line['bu'	] 	= bu_name
			else:
				new_itd_stock_line['bu'] = '-'

			for reservation_line in reserv_list:
				if (reservation_line['material'].lower() == new_itd_stock_line['material'].lower()):
					new_itd_stock_line['stock_4000'	] 	-= reservation_line['res_4000_qty']
					new_itd_stock_line['stock_3400'	] 	-= reservation_line['res_3400_qty']

			if new_itd_stock_line['stock_4000'] < 0 : new_itd_stock_line['stock_4000'] = 0.0
			if new_itd_stock_line['stock_3400'] < 0 : new_itd_stock_line['stock_3400'] = 0.0

			#working with 1C information
			new_itd_stock_line['out_of_stock'] = 0.0
			new_itd_stock_line['refurbished'] = 0.0

			for material in onec_list:
				if (new_itd_stock_line['material'].lower() == material['material'].lower()):
					if material['status'] == 'out of stock':
						new_itd_stock_line['out_of_stock'] = material['qty']
					elif material['status'] == 'refurbished':
						new_itd_stock_line['refurbished'] = material['qty']

				if ((new_itd_stock_line['material'].lower() + 'q' == material['material'].lower()) and
						(material['status'] == 'refurbished')):
					#print(new_itd_stock_line['material'])
					new_itd_stock_line['refurbished'] += material['qty']

			new_itd_stock_list.append(new_itd_stock_line)

		new_itd_stock_list.sort( key = lambda item: ( item['material'] ) )

		return new_itd_stock_list

	except Exception as e:
		elog('gen_new_stock_list','{d}_{type}\n{trb1}\n{trb2} '.format( d=datetime.datetime.now(),
																		type=str(type(e)),
																		trb1=traceback.format_tb(sys.exc_info()[2])[0],
																		trb2=str(sys.exc_info()[1])))
		return []

#- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
#- - - - - N E W   S T O C K   F I L E - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
#- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
	
def new_stock_file_make():
	
	try:
		#print ('new_stock_file_make start time :', datetime.datetime.now())
		
		wb_new 		= Workbook()
		ws_1 		= wb_new.active
		ws_1.title	= 'Stock'
		ws_2		= wb_new.create_sheet('Cross reference ITD')

		ws_1.sheet_view.showGridLines = False
		ws_2.sheet_view.showGridLines = False

		# - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
		# WS 1
		# - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -

		for row_index in range(1, 10) :
			ws_1['{}1'.format(get_column_letter(row_index)) ].alignment = Alignment( horizontal='center',
																					 vertical='center',
																				 	 wrap_text = True )

		ws_1.column_dimensions['A'].width 	= 25
		ws_1.column_dimensions['B'].width 	= 45
		ws_1.column_dimensions['C'].width 	= 45
		ws_1.column_dimensions['D'].width 	= 17
		ws_1.column_dimensions['E'].width 	= 17
		ws_1.column_dimensions['F'].width 	= 20
		ws_1.column_dimensions['G'].width 	= 20
		ws_1.column_dimensions['H'].width 	= 25
		ws_1.column_dimensions['I'].width 	= 20

		ws_1['A1'].value 	= 'Material'
		ws_1['B1'].value 	= 'Description'
		ws_1['C1'].value 	= 'Material Status'
		ws_1['D1'].value 	= 'Free Stock'
		ws_1['E1'].value 	= 'Contract Stock'
		ws_1['F1'].value 	= 'Refurbished'
		ws_1['G1'].value 	= 'Not On Balance'
		ws_1['H1'].value 	= 'Business Unit'
		ws_1['I1'].value 	= 'Delivery time'


		#print ('new_itd_list end time :', datetime.datetime.now() )
		
		i = 2
		#new_itd_stock_list.sort( reverse = True, key = lambda item: ( item['lenth'] ) ) #sort by 'lenth','time_start'
			
		for line in new_itd_stock_list:
			ws_1.cell(row=i,column=1).value = line['material'	]
			ws_1.cell(row=i,column=2).value = line['description']
			ws_1.cell(row=i,column=3).value = line['status_name']
			ws_1.cell(row=i,column=4).value = line['stock_4000'	]
			ws_1.cell(row=i,column=5).value = line['stock_3400'	]
			ws_1.cell(row=i,column=6).value = line['refurbished']
			ws_1.cell(row=i,column=7).value = line['out_of_stock']
			ws_1.cell(row=i,column=8).value = line['bu']
			ws_1.cell(row=i,column=9).value = line['pdt']

			i = i + 1
		
		style 					= TableStyleInfo(	name = "TableStyleMedium2",
													showRowStripes = True,
												   	showColumnStripes = False)

		table_1 				= Table( displayName = u"Table_1", ref = "$A$1:$I${index}".format(index = i - 1 ) )
		table_1.tableStyleInfo 	= style
		ws_1.add_table(table_1)

		# - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
		# WS 2
		# - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -

		for row_index in range(1, 6):
			ws_2['{}1'.format(get_column_letter(row_index))].alignment = Alignment(horizontal='center',
																				   vertical='center',
																				   wrap_text=True)

		ws_2.column_dimensions['A'].width = 25
		ws_2.column_dimensions['B'].width = 25
		ws_2.column_dimensions['C'].width = 25
		ws_2.column_dimensions['D'].width = 25


		ws_2['A1'].value = 'Material'
		ws_2['B1'].value = 'Material2'
		ws_2['C1'].value = 'Material3'
		ws_2['D1'].value = 'Material4'

		# print ('new_itd_list end time :', datetime.datetime.now())

		i = 2

		last_itd_cross_list.sort(key=lambda item: (item['material1']))
		last_itd_cross_list.sort( reverse = True, key = lambda item: ( item['lenth'] ) ) #sort by 'lenth'

		for line in last_itd_cross_list:
			ws_2.cell(row=i, column=1).value = line['material1']
			ws_2.cell(row=i, column=2).value = line['material2']
			ws_2.cell(row=i, column=3).value = line['material3']
			ws_2.cell(row=i, column=4).value = line['material4']

			i = i + 1

		table_2 = Table(displayName=u"Table_2", ref="$A$1:$D${index}".format(index=i - 1))
		table_2.tableStyleInfo = style
		ws_2.add_table(table_2)

		now 					= datetime.datetime.now()
		#date 					= '{data}'.format(data = date.today)
		date 					= '{y}.{m}.{d}'.format(y = now.year, m = now.month, d = now.day)
		time 					= '{h}.{m}.{s}'.format(h = now.hour, m = now.minute, s = now.second)
		wb_new.save('{path}\\results\\Stock_Update_{date}_{time}.xlsx'.format( path = path,
																					 date = date,
																					 time = time ))

	except Exception as e:
		elog('new_stock_file_make','{d}_{type}\n{trb1}\n{trb2} '.format( d = datetime.datetime.now(),
																	   type=str(type(e)),
																	   trb1=traceback.format_tb(sys.exc_info()[2])[0],
																	   trb2=str(sys.exc_info()[1])))

# - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
# - - - - - N E W   P A R T N E R   F I L E - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
# - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -

def new_partner_file_make():
	try:
		wb_new = Workbook()
		ws_1 = wb_new.active
		ws_1.title = 'Stock'
		ws_2 = wb_new.create_sheet('Cross reference ITD')

		ws_1.sheet_view.showGridLines = False
		ws_2.sheet_view.showGridLines = False

		# - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
		# WS 1
		# - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -

		for row_index in range(1, 7):
			ws_1['{}1'.format(get_column_letter(row_index))].alignment = Alignment(horizontal='center',
																				   vertical='center',
																				   wrap_text=True)

		ws_1.column_dimensions['A'].width = 25
		ws_1.column_dimensions['B'].width = 45
		ws_1.column_dimensions['C'].width = 45
		ws_1.column_dimensions['D'].width = 20
		ws_1.column_dimensions['E'].width = 20
		ws_1.column_dimensions['F'].width = 25

		ws_1['A1'].value = 'Material'
		ws_1['B1'].value = 'Description'
		ws_1['C1'].value = 'Material Status'
		ws_1['D1'].value = 'Free Stock'
		ws_1['E1'].value = 'Contract Stock'
		ws_1['F1'].value = 'Business Unit'

		# print ('new_itd_list end time :', datetime.datetime.now())

		i = 2
		new_itd_stock_list.sort(key=lambda item: (item['material']))

		for line in new_itd_stock_list:
			ws_1.cell(row=i, column=1).value = line['material']
			ws_1.cell(row=i, column=2).value = line['description']
			ws_1.cell(row=i, column=3).value = line['status_name']
			ws_1.cell(row=i, column=4).value = line['stock_4000']
			ws_1.cell(row=i, column=5).value = line['stock_3400']
			ws_1.cell(row=i, column=6).value = line['bu']

			i = i + 1

		style = TableStyleInfo(name="TableStyleMedium2",
							   showRowStripes=True,
							   showColumnStripes=False)

		table_1 = Table(displayName=u"Table_1", ref="$A$1:$F${index}".format(index=i - 1))
		table_1.tableStyleInfo = style
		ws_1.add_table(table_1)

		# - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
		# WS 2
		# - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -

		for row_index in range(1, 6):
			ws_2['{}1'.format(get_column_letter(row_index))].alignment = Alignment(horizontal='center',
																				   vertical='center',
																				   wrap_text=True)

		ws_2.column_dimensions['A'].width = 25
		ws_2.column_dimensions['B'].width = 25
		ws_2.column_dimensions['C'].width = 25
		ws_2.column_dimensions['D'].width = 25

		ws_2['A1'].value = 'Material'
		ws_2['B1'].value = 'Material2'
		ws_2['C1'].value = 'Material3'
		ws_2['D1'].value = 'Material4'

		i = 2

		last_itd_cross_list.sort(key=lambda item: (item['material1']))
		last_itd_cross_list.sort(reverse=True, key=lambda item: (item['lenth']))  # sort by 'lenth'

		for line in last_itd_cross_list:
			ws_2.cell(row=i, column=1).value = line['material1']
			ws_2.cell(row=i, column=2).value = line['material2']
			ws_2.cell(row=i, column=3).value = line['material3']
			ws_2.cell(row=i, column=4).value = line['material4']

			i = i + 1

		table_2 = Table(displayName=u"Table_2", ref="$A$1:$D${index}".format(index=i - 1))
		table_2.tableStyleInfo = style
		ws_2.add_table(table_2)

		now = datetime.datetime.now()
		date = '{y}.{m}.{d}'.format(y=now.year, m=now.month, d=now.day)
		time = '{h}.{m}.{s}'.format(h=now.hour, m=now.minute, s=now.second)
		wb_new.save('{path}\\partners_results\\Stock_Update_{date}.xlsx'.format(path=path,
																				date=date))

	except Exception as e:
		elog('new_partner_file_make', '{d}_{type}\n{trb1}\n{trb2} '.format(d=datetime.datetime.now(),
																		 type=str(type(e)),
																		 trb1=traceback.format_tb(sys.exc_info()[2])[0],
																		 trb2=str(sys.exc_info()[1])))

#- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
#- - - - - L A U N C H I N G - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
#- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -

start =  datetime.datetime.now()
print ('reserv_loading')
print  (datetime.datetime.now())
reserv_list = reserv_loading()
print ('stock_loading')
print  (datetime.datetime.now())
stock_list 	= stock_loading()
print ('last_itd_loading')
print  (datetime.datetime.now())
last_itd_stock_list, last_itd_cross_list = last_itd_loading()
print ('onec_loading')
print  (datetime.datetime.now())
onec_file = onec_loading()
print ('gen_new_stock_list')
print  (datetime.datetime.now())
new_itd_stock_list = gen_new_stock_list()
print ('new_stock_file_make')
print  (datetime.datetime.now())
new_stock_file_make()
print ('new_partner_file_make')
print  (datetime.datetime.now())

end = datetime.datetime.now()
timedelta = end - start

print('Time delta_final: {}'.format(timedelta))

